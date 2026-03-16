# interface.py

import dearpygui.dearpygui as dpg
import tkinter as tk
from tkinter import filedialog
from classes import *
import time
from openpyxl import Workbook

db = ''
db_2 = ''

step = 10000


def set_ui_enabled(enabled: bool):
    pass

def load_to_db(sender):
    try:
        global db

        set_ui_enabled(False)
        if sender == 'btn_table':
            root = tk.Tk()
            root.withdraw()
            file_path = filedialog.askdirectory()
            dpg.configure_item(item='input_dir', default_value=file_path)

            db = DateBase(db_name='input_datas.db', exists=False)  # создаем базу данных

            input_table_header = [
                'Timestamp_UTC',
                'Ua',
                'Ua_phase',
                'Ub',
                'Ub_phase',
                'Uc',
                'Uc_phase',
                'Ia',
                'Ia_phase',
                'Ib',
                'Ib_phase',
                'Ic',
                'Ic_phase',
                'f',
                'pribor_num'
            ]

            # вводим данные в бд
            if db.select_table("SELECT name FROM sqlite_master WHERE type='table' and name='input_data';") == []:

                db.create_db_table(  # создаем таблицу для входных данных
                    'input_data',
                    input_table_header
                )

                all_tables = find_files(file_path)  # находим пути всех таблиц

                for table in all_tables:  # заносим все таблицы в базу данных
                    file = DataTable(table)
                    # print('name', file.file_name, table)
                    time_1 = time.time()
                    db.insert_db_table('input_data', input_table_header,
                                       format_data(file.data, [0, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 31],
                                                   table[-10:-8]))
                    time_2 = time.time()
                    v = 1 / (time_2 - time_1)
                    t = (len(all_tables) - all_tables.index(table) - 1) / v
                    percent = round((all_tables.index(table) + 1) / len(all_tables), 2)

                    dpg.configure_item(item='progress_bar', show=True, default_value=percent)
                    dpg.configure_item(item='text_1', default_value=f'(1/2) {percent * 100} %, осталось {nice_t(t)}')

                hv_lv_table_header = [
                    'Timestamp_UTC',
                    'UA_hv',
                    'UA_hv_phase',
                    'UB_hv',
                    'UB_hv_phase',
                    'UC_hv',
                    'UC_hv_phase',
                    'IA_hv',
                    'IA_hv_phase',
                    'IB_hv',
                    'IB_hv_phase',
                    'IC_hv',
                    'IC_hv_phase',
                    'f_hv',
                    'pribor_num_hv',
                    'Timestamp_UTC_lv',
                    'UA_lv',
                    'UA_lv_phase',
                    'UB_lv',
                    'UB_lv_phase',
                    'UC_lv',
                    'UC_lv_phase',
                    'IA_lv',
                    'IA_lv_phase',
                    'IB_lv',
                    'IB_lv_phase',
                    'IC_lv',
                    'IC_lv_phase',
                    'f_lv',
                    'pribor_num_lv',
                    'temp'
                ]

                db.create_db_table(  # создаем таблицу для данных вн-нн
                    'hv_lv_data',
                    hv_lv_table_header
                )
                # заносим hv-lv в бД
                num_of_timestamp = db.select_table("""SELECT COUNT(a.Timestamp_UTC)
                                                                FROM input_data as a
                                                                INNER JOIN input_data as b ON a.Timestamp_UTC = b.Timestamp_UTC
                                                                WHERE a.pribor_num LIKE '74' AND b.pribor_num LIKE '76';""")[0][0]

                db.query("""INSERT INTO hv_lv_data SELECT a.*, b.*, '-'
                                            FROM input_data as a
                                            INNER JOIN input_data as b ON a.Timestamp_UTC = b.Timestamp_UTC
                                            WHERE a.pribor_num LIKE '74' AND b.pribor_num LIKE '76';""")

                for i in range(0, num_of_timestamp, step):
                    time_1 = time.time()
                    time.sleep(0.001)
                    time_2 = time.time()
                    v = step / (time_2 - time_1)
                    t = (num_of_timestamp - i) / v
                    percent = round((i + step) / num_of_timestamp, 2)

                    dpg.configure_item(item='progress_bar', show=True, default_value=percent)
                    dpg.configure_item(item='text_1', default_value=f'(2/2) {percent * 100} %, осталось {nice_t(t)}')

                dpg.configure_item(item='text_1', default_value='Данные загружены в базу данных')
                dpg.configure_item(item='progress_bar', show=False, default_value=0)

        elif sender == 'btn_db':

            root = tk.Tk()
            root.withdraw()
            file_path = filedialog.askopenfilename(filetypes=[("Data base files", ".db")])
            dpg.configure_item(item='input_dir', default_value=file_path)

            if dpg.get_value(item='input_dir'):
                db = DateBase(db_name=file_path, exists=True)  # создаем базу данных
                dpg.configure_item(item='text_1', default_value='Файл базы данных загружен')
            else:
                dpg.set_value(item='text_1', value='')
    except Exception as e:
        show_error_dialog(f"Ошибка при загрузке данных: {str(e)}")
    finally:
        set_ui_enabled(True)

def calculate():
    try:
        global db, db_2

        ki_lv = float(dpg.get_value(item='kt nn').replace(',', '.'))
        ku_hv = float(dpg.get_value(item='kn vn').replace(',', '.'))
        ki_hv = float(dpg.get_value(item='kt vn').replace(',', '.'))
        ku_lv = float(dpg.get_value(item='kn nn').replace(',', '.'))
        k = float(dpg.get_value(item='kt').replace(',', '.'))
        Snom = float(dpg.get_value(item='S').replace(',', '.'))
        kt = float(dpg.get_value(item='kt').replace(',', '.'))
        Rt = float(dpg.get_value(item='Rt').replace(',', '.'))
        Xt = float(dpg.get_value(item='Xt').replace(',', '.'))
        Gt = float(dpg.get_value(item='Gt').replace(',', '.'))
        Bt = float(dpg.get_value(item='Bt').replace(',', '.'))


        kn_vn_a = float(dpg.get_value(item='kn vn a').replace(',', '.'))
        kn_vn_w = float(dpg.get_value(item='kn vn w').replace(',', '.'))

        koff = dpg.get_value(item='koff')

        count = db.select_table(f"""SELECT COUNT(Timestamp_UTC)
                                    FROM hv_lv_data
                                    """)[0][0]

        bd_2_heading = [
            'Timestamp_UTC',
            'f_hv',
            'f_lv',
            'UA_hv_value',
            'UA_hv_phase',
            'UB_hv_value',
            'UB_hv_phase',
            'UC_hv_value',
            'UC_hv_phase',
            'IA_hv_value',
            'IA_hv_phase',
            'IB_hv_value',
            'IB_hv_phase',
            'IC_hv_value',
            'IC_hv_phase',
            'UA_lv_value',
            'UA_lv_phase',
            'UB_lv_value',
            'UB_lv_phase',
            'UC_lv_value',
            'UC_lv_phase',
            'IA_lv_value',
            'IA_lv_phase',
            'IB_lv_value',
            'IB_lv_phase',
            'IC_lv_value',
            'IC_lv_phase',
            'U1_hv_value',
            'U1_hv_phase',
            'I1_hv_value',
            'I1_hv_phase',
            'U1_lv_value',
            'U1_lv_phase',
            'I1_lv_value',
            'I1_lv_phase',
            'I_m',
            'I_m_phase',
            'U0',
            'U0_phase',
            'R_T',
            'X_T',
            'G_microS_T',
            'B_microS_T',
            'R_G',
            'X_G',
            'G_microS_G',
            'B_microS_G',
            'S_hv_total',
            'P_hv',
            'Q_hv',
            'S_lv_total',
            'P_lv',
            'Q_lv',
            'dS_total',
            'dP',
            'dQ',
            'Kz_percent',
            'Kt',
            'Kpd',

            'S_kz_T',
            'P_kz_T',
            'Q_kz_T',

            'S_hh_T',
            'P_hh_T',
            'Q_hh_T',

            'S_kz_G',
            'P_kz_G',
            'Q_kz_G',

            'S_hh_G',
            'P_hh_G',
            'Q_hh_G',

            'bR_T',
            'bX_T',
            'bG_T',
            'bB_T',

            'bR_G',
            'bX_G',
            'bG_G',
            'bB_G',

            'temp'
        ]

        db_2 = DateBase(db_name='calculate_data.db')

        db_2.create_db_table('all_values', bd_2_heading)

        for i in range(0, count, step):
            time_1 = time.time()

            # берем данные из таблицы за нужный интервал
            table_1 = db.select_table(
                f"""SELECT * FROM hv_lv_data LIMIT {i}, {step};""")
            DbTable_1 = DbTable(table_1)

            # конвертируем данные в классы
            DbTable_1.convert_class([
                (str, [0]),  # Timestamp_UTC_hv
                (ComplexValue, [1, 2, 13,]),  # UA_hv
                (ComplexValue, [3, 4, 13]),  # UB_hv
                (ComplexValue, [5, 6, 13]),  # UC_hv
                (ComplexValue, [7, 8, 13]),  # IA_hv
                (ComplexValue, [9, 10, 13]),  # IB_hv
                (ComplexValue, [11, 12, 13]),  # IC_hv
                (str, [13]),  # f_hv
                (ComplexValue, [16, 17, 28]),  # UA_lv (без коэффициента)
                (ComplexValue, [18, 19, 28]),  # UB_lv  (без коэффициента)
                (ComplexValue, [20, 21, 28]),  # UC_lv  (без коэффициента)
                (ComplexValue, [22, 23, 28]),  # IA_lv
                (ComplexValue, [24, 25, 28]),  # IB_lv
                (ComplexValue, [26, 27, 28]),  # IC_lv
                (str, [28]),  # f_lv
                (float, [1])
            ])

            # рассчитываем данные
            DbTable_1.add_calculate_values(
                Snom, kt, Rt, Xt, Gt, Bt,
                kn_vn_a, kn_vn_w,
                koff,
                ku_hv,
                ki_hv,
                ki_lv
            )

            # конвертируем для базы данных
            DbTable_1.convert_str()

            db_2.insert_db_table('all_values', bd_2_heading, DbTable_1.str_data)

            time_2 = time.time()
            v = step / (time_2 - time_1)
            t = (count - i) / v
            percent = round((i + step) / count, 2)

            dpg.configure_item(item='progress_bar_2', show=True, default_value=percent)
            dpg.configure_item(item='text_2', default_value=f'{percent * 100} %, осталось {nice_t(t)}')

        dpg.configure_item(item='progress_bar_2', show=False, default_value=0)
        dpg.configure_item('text_2', default_value='Вычисления загружены в базу данных')
        dpg.configure_item('add_lin_graph_btn', show=True)
    except Exception as e:
        show_error_dialog(f"Ошибка при расчете: {str(e)}")
def calc_trans():
    try:
        katalog_1 = KatalogData(
            float(dpg.get_value(item='S').replace(',', '.')),
            float(dpg.get_value(item='U vn').replace(',', '.')),
            float(dpg.get_value(item='U nn').replace(',', '.')),
            float(dpg.get_value(item='U k').replace(',', '.')),
            float(dpg.get_value(item='dPk').replace(',', '.')),
            float(dpg.get_value(item='Phh').replace(',', '.')),
            float(dpg.get_value(item='Ihh').replace(',', '.'))
        )

        dpg.configure_item('dQ', default_value=str(round(katalog_1.dQh, 4)).replace('.', ','))
        dpg.configure_item('Rt', default_value=str(round(katalog_1.Rt, 4)).replace('.', ','))
        dpg.configure_item('Xt', default_value=str(round(katalog_1.Xt, 4)).replace('.', ','))
        dpg.configure_item('Gt', default_value=str(round(katalog_1.Gt, 4)).replace('.', ','))
        dpg.configure_item('Bt', default_value=str(round(katalog_1.Bt, 4)).replace('.', ','))
        dpg.configure_item('kt', default_value=str(round(katalog_1.Kt, 4)).replace('.', ','))
    except Exception as e:
        show_error_dialog(f"Ошибка расчета трансформатора: {str(e)}")

def load_calc_file():
    try:
        global db_2

        root = tk.Tk()
        root.withdraw()
        file_path = filedialog.askopenfilename(filetypes=[("Data base files", ".db")])

        if file_path:
            db_2 = DateBase(db_name=file_path, exists=True)  # создаем базу данных
            dpg.configure_item(item='text_2', default_value='Файл базы данных вычислений загружен')

            dpg.configure_item('add_lin_graph_btn', show=True)
    except Exception as e:
        show_error_dialog(f"Ошибка загрузки файла: {str(e)}")


lin_fields = []
dot_fields = []

lin_graphs = []
lin_series = {}
lin_axises = {}

items = {
    # Таймстамп и частота
    't': 'Timestamp_UTC',
    'f вн, кГц': 'f_hv',
    'f нн, кГц': 'f_lv',

    # Высокое напряжение (вн)
    'UA вн, В': 'UA_hv_value',
    'UA вн, фаза, °': 'UA_hv_phase',
    'UB вн, В': 'UB_hv_value',
    'UB вн, фаза, °': 'UB_hv_phase',
    'UC вн, В': 'UC_hv_value',
    'UC вн, фаза, °': 'UC_hv_phase',

    # Токи высокого напряжения (вн)
    'IA вн, А': 'IA_hv_value',
    'IA вн, фаза, °': 'IA_hv_phase',
    'IB вн, А': 'IB_hv_value',
    'IB вн, фаза, °': 'IB_hv_phase',
    'IC вн, А': 'IC_hv_value',
    'IC вн, фаза, °': 'IC_hv_phase',

    # Низкое напряжение (нн)
    'UA нн, В': 'UA_lv_value',
    'UA нн, фаза, °': 'UA_lv_phase',
    'UB нн, В': 'UB_lv_value',
    'UB нн, фаза, °': 'UB_lv_phase',
    'UC нн, В': 'UC_lv_value',
    'UC нн, фаза, °': 'UC_lv_phase',

    # Токи низкого напряжения (нн)
    'IA нн, А': 'IA_lv_value',
    'IA нн, фаза, °': 'IA_lv_phase',
    'IB нн, А': 'IB_lv_value',
    'IB нн, фаза, °': 'IB_lv_phase',
    'IC нн, А': 'IC_lv_value',
    'IC нн, фаза, °': 'IC_lv_phase',

    # Симметричные составляющие
    'U1 вн, В': 'U1_hv_value',
    'U1 вн, фаза, °': 'U1_hv_phase',
    'I1 вн, А': 'I1_hv_value',
    'I1 вн, фаза, °': 'I1_hv_phase',
    'U1 нн, В': 'U1_lv_value',
    'U1 нн, фаза, °': 'U1_lv_phase',
    'I1 нн, А': 'I1_lv_value',
    'I1 нн, фаза, °': 'I1_lv_phase',

    # Мешающие параметры
    'Iм, А': 'I_m',
    'Iм, фаза, °': 'I_m_phase',
    'U0, В': 'U0',
    'U0, фаза, °': 'U0_phase',

    # Параметры сети
    'R(Т), Ом': 'R_T',
    'X(Т), Ом': 'X_T',
    'G(Т), мкСм': 'G_microS_T',
    'B(Т), мкСм': 'B_microS_T',
    'R(Г), Ом': 'R_G',
    'X(Г), Ом': 'X_G',
    'G(Г), мкСм': 'G_microS_G',
    'B(Г), мкСм': 'B_microS_G',

    # Мощности
    'Sвн, ВА': 'S_hv_total',
    'Pвн, Вт': 'P_hv',
    'Qвн, вар': 'Q_hv',
    'Sнн, ВА': 'S_lv_total',
    'Pнн, Вт': 'P_lv',
    'Qнн, вар': 'Q_lv',

    # Дельта-параметры
    'dS, ВА': 'dS_total',
    'dP, Вт': 'dP',
    'dQ, вар': 'dQ',

    # Коэффициенты
    'Kз, %': 'Kz_percent',
    'Kт': 'Kt',
    'KПД, %': 'Kpd',

    'dSкз(Т), ВА': 'S_kz_T',
    'dPкз(Т), Вт': 'P_kz_T',
    'dQкз(Т), вар': 'Q_kz_T',

    'dSхх(Т), ВА': 'S_hh_G',
    'dPхх(Т), Вт': 'P_hh_G',
    'dQхх(Т), вар': 'Q_hh_G',

    'dSкз(Г), ВА': 'S_kz_G',
    'dPкз(Г), Вт': 'P_kz_G',
    'dQкз(Г), вар': 'Q_kz_G',

    'dSхх(Г), ВА': 'S_hh_G',
    'dPхх(Г), Вт': 'P_hh_G',
    'dQхх(Г), вар': 'Q_hh_G',

    'bR(Т), %': 'bR_T',
    'bX(Т), %': 'bX_T',
    'bG(Т), %': 'bG_T',
    'bB(Т), %': 'bB_T',

    'bR(Г), %': 'bR_G',
    'bX(Г), %': 'bX_G',
    'bG(Г), %': 'bG_G',
    'bB(Г), %': 'bB_G',
}

def comma_input_callback(sender, app_data):
    if isinstance(app_data, str):
        new_value = app_data.replace('.', ',')
        if new_value != app_data:
            dpg.set_value(sender, new_value)

def show_error_dialog(message):
    with dpg.window(tag="error_dialog", label="Ошибка", modal=True, show=True, width=400, height=150):
        dpg.add_text(message)
        dpg.add_button(label="OK", width=75, callback=lambda: dpg.delete_item("error_dialog"))

def add_lin_plot():
    global lin_fields, lin_axises



    if len(lin_fields) == 0:
        i = 1
    else:
        i = lin_fields[-1] + 1


    with dpg.tab(parent='tab_bar1', tag=f'lin_field_{i}', label=f'Поле {i}'):
        lin_fields.append(i)

        with dpg.plot(tag=f'lin_field_{i}-plot', width=-1, height=400):
            x_axis = dpg.add_plot_axis(dpg.mvXAxis, label="X", tag=f"axis_x_{i}")
            y_axis = dpg.add_plot_axis(dpg.mvYAxis, label="Y", tag=f"axis_y_{i}")
            dpg.add_plot_legend()

        lin_axises[i] = [x_axis, y_axis]


        dpg.add_button(label='+', tag=f'lin_field_{i}_add_graph_btn', callback=add_lin_graph, user_data=(i, y_axis))
        dpg.add_button(label='Удалить поле', tag=f'lin_field_{i}_del_field', callback=del_field, user_data=i)

def add_lin_graph(sender, app_data, user_data):
    global lin_graphs, lin_series

    i, y_axis= user_data # номер поля
    if i not in lin_series:
        lin_series[i] = {}

    if len(lin_graphs) == 0:
        j = 1
    else:
        j = lin_graphs[-1][1] + 1

    lin_graphs.append([i, j])

    with dpg.group(horizontal=True, parent=f'lin_field_{i}', tag=f'lin_graph_{i}_{j}'):
        dpg.add_button(label='-', callback=del_lin_graph, user_data=[i, j])
        dpg.add_text(default_value='y =')
        dpg.add_combo(items=list(items.keys()), width=150, tag=f'combo_y_{i}_{j}')
        dpg.add_text(default_value='x =')
        dpg.add_combo(items=list(items.keys()), width=150, tag=f'combo_x_{i}_{j}')
        dpg.add_text(default_value='t1 =')
        dpg.add_input_text(default_value='05/06/2024 00:00:00.000', width=200, tag=f'time_1_{i}_{j}', callback=comma_input_callback)
        dpg.add_text(default_value='t2 =')
        dpg.add_input_text(default_value='05/06/2024 00:01:00.000', width=200, tag=f'time_2_{i}_{j}', callback=comma_input_callback)
        dpg.add_text(default_value='dt, с =')
        dpg.add_input_text(default_value='0.02', width=100, tag=f'dt_{i}_{j}', callback=comma_input_callback)
        dpg.add_button(label='+', callback=build_lin_graph, user_data=[i, j, y_axis])
        dpg.add_button(label='zoom', callback=zoom_graph, user_data=[i, j])
        dpg.add_loading_indicator(show=False, tag=f'load_{i}_{j}', radius=1, thickness=2)
        dpg.add_button(label='Excel', callback=save_to_excel, user_data=[i, j])

    lin_series[i][j] = None

def build_lin_graph(sender, app_data, user_data):
    # try:
    global db_2, lin_series

    i, j, y_axis = user_data

    dpg.configure_item(item=f'load_{i}_{j}', show=True)

    if lin_series[i][j] is not None:
        if dpg.does_item_exist(lin_series[i][j]):
            dpg.delete_item(lin_series[i][j])
        lin_series[i][j] = None

    values = {
            'y': dpg.get_value(f'combo_y_{i}_{j}'),
            'x': dpg.get_value(f'combo_x_{i}_{j}'),
            'time_1': dpg.get_value(f'time_1_{i}_{j}'),
            'time_2': dpg.get_value(f'time_2_{i}_{j}'),
            'dt': dpg.get_value(f'dt_{i}_{j}')
        }
    print(f'Поле {i}, график {j}')

    graph_limit = 10000



    count_dots = db_2.select_table(f"SELECT COUNT({items[values['y']]}) FROM all_values WHERE Timestamp_UTC >= '{values['time_1']}' AND Timestamp_UTC < '{values['time_2']}'")[0][0]
    print(f'количество точек всего {count_dots}')

    if count_dots > graph_limit:
        dt = float(((count_dots // graph_limit) * 0.02))
        if float(values['dt']) < dt:
            dpg.configure_item(item=f'dt_{i}_{j}', default_value=dt)
            values['dt'] = dt

    if float(values['dt']) / 0.02 > graph_limit:
        dpg.configure_item(item=f'dt_{i}_{j}', default_value=graph_limit * 0.02)
        values['dt'] = graph_limit * 0.02

    print(f'количество точек всего {count_dots}, надо усреднять по {values["dt"]} с')


    table_for_graph_y = []
    table_for_graph_x = []

    time_label_x = []
    time_label_y = []

    for k in range(0, count_dots, graph_limit):
        print(f'пакет № k = {k}')

        graph_data = DbTable(db_2.select_table(f"SELECT {items[values['y']]}, {items[values['x']]} FROM all_values WHERE Timestamp_UTC >= '{values['time_1']}' AND Timestamp_UTC < '{values['time_2']}' LIMIT {k}, {graph_limit};"))

        usred_data_x = graph_data.usred(values['dt'])[1]
        usred_data_y = graph_data.usred(values['dt'])[0]

        if values['x'] == 't':
            time_label_x.extend(usred_data_x)
        else:
            table_for_graph_x.extend(usred_data_x)
        if values['y'] == 't':
            time_label_y.extend(usred_data_y)
        else:
            table_for_graph_y.extend(usred_data_y)
    if values['x'] == 't':
        table_for_graph_x.extend(range(0, len(time_label_x)))
    if values['y'] == 't':
        table_for_graph_y.extend(range(0, len(time_label_y)))

    print(f'длина списка x {len(table_for_graph_y)}')
    print(table_for_graph_y, time_label_y, table_for_graph_x, time_label_x)

    series_id = dpg.add_scatter_series(y=table_for_graph_y, x=table_for_graph_x, parent=y_axis, tag=f'series_{i}_{j}', label=f'{values["y"]} = f({values["x"]})')
    print(f'{values["y"]} = f({values["x"]})')

    print(tuple(zip(table_for_graph_y, time_label_y)))
    lin_series[i][j] = series_id
    zoom_graph(None, None, [i, j])

    dpg.configure_item(item=f'load_{i}_{j}', show=False)

    graph_data = {
        'x': table_for_graph_x,
        'y': table_for_graph_y,
        'time_labels_x': time_label_x,
        'time_labels_y': time_label_y,
        'x_label': dpg.get_value(f'combo_x_{i}_{j}'),
        'y_label': dpg.get_value(f'combo_y_{i}_{j}')
    }
    dpg.set_item_user_data(f'lin_graph_{i}_{j}', graph_data)

    # except Exception as e:
    #     show_error_dialog(f"Ошибка построения графика: {str(e)}")
    #     print
    #     # Скрываем индикатор загрузки при ошибке
    #     i, j = user_data[:2]  # Берем первые два элемента
    #     dpg.configure_item(item=f'load_{i}_{j}', show=False)

# Добавляем функцию приближения
def zoom_graph(sender, app_data, user_data):
    global lin_series
    i, j = user_data

    print(i, j, lin_series)

    # Получаем ID графика
    series_id = lin_series[i][j]
    if not series_id or not dpg.does_item_exist(series_id):
        return

    # Получаем данные графика
    data = dpg.get_value(series_id)
    if not data or len(data[0]) == 0 or len(data[1]) == 0:
        return

    # Находим min/max по осям
    min_x, max_x = min(data[0]), max(data[0])
    min_y, max_y = min(data[1]), max(data[1])

    # Добавляем небольшой отступ (10% от диапазона)
    x_padding = (max_x - min_x) * 0.1
    y_padding = (max_y - min_y) * 0.1

    print(lin_axises)

    # Устанавливаем новые границы осей
    dpg.set_axis_limits(lin_axises[i][0], min_x, max_x)
    dpg.set_axis_limits(lin_axises[i][1], min_y, max_y)

    time.sleep(0.1)

    dpg.set_axis_limits_auto(lin_axises[i][0])
    dpg.set_axis_limits_auto(lin_axises[i][1])

def save_to_excel(sender, app_data, user_data):
    try:
        i, j = user_data
        graph_data = dpg.get_item_user_data(f'lin_graph_{i}_{j}')

        dpg.configure_item(item=f'load_{i}_{j}', show=True)

        if not graph_data or not graph_data.get('x') or not graph_data.get('y'):
            dpg.configure_item(item='text_2', default_value='Нет данных для сохранения')
            return

        # Получаем параметры графика
        time_1 = dpg.get_value(f'time_1_{i}_{j}')
        time_2 = dpg.get_value(f'time_2_{i}_{j}')
        dt = dpg.get_value(f'dt_{i}_{j}')

        # Запрашиваем путь для сохранения
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Сохранить данные графика"
        )

        if not file_path:  # Пользователь отменил сохранение
            return

        try:
            # Создаем новую книгу Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Graph Data"

            # Добавляем метаданные в левую часть
            metadata = [
                ("Период:", f"с {time_1} по {time_2}"),
                ("Интервал усреднения:", f"{dt} с")
            ]

            # Записываем метаданные в столбец A
            for row_idx, (key, value) in enumerate(metadata, start=1):
                ws.cell(row=row_idx, column=1, value=key)
                ws.cell(row=row_idx, column=2, value=value)

            # Добавляем заголовки данных в столбцы C и D
            ws.cell(row=1, column=3, value=graph_data['x_label'])
            ws.cell(row=1, column=4, value=graph_data['y_label'])

            # Данные
            for idx, (x_val, y_val) in enumerate(zip(graph_data['x'], graph_data['y']), start=2):
                # print(len(enumerate(zip(graph_data['x'], graph_data['y']))))

                if len(graph_data['time_labels_x']) > 0:

                    ws.cell(row=idx, column=3, value=graph_data['time_labels_x'][x_val-2])
                else:
                    print(len(graph_data))
                    ws.cell(row=idx, column=3, value=x_val)
                if len(graph_data['time_labels_y']) > 0:
                    ws.cell(row=idx, column=4, value=graph_data['time_labels_y'][y_val-2])
                else:
                    ws.cell(row=idx, column=4, value=y_val)
            # Автонастройка ширины столбцов
            for col in ['A', 'B', 'C', 'D']:
                max_length = 0
                for row in range(1, ws.max_row + 1):
                    cell_value = ws[f"{col}{row}"].value
                    if cell_value and len(str(cell_value)) > max_length:
                        max_length = len(str(cell_value))
                ws.column_dimensions[col].width = max_length + 2

            # Сохраняем файл
            wb.save(file_path)
            # dpg.configure_item(item='text_2', default_value=f'Данные сохранены в {file_path}')

        except Exception as e:
            print(e)
            # dpg.configure_item(item='text_2', default_value=f'Ошибка сохранения: {str(e)}')

        dpg.configure_item(item=f'load_{i}_{j}', show=False)
    except Exception as e:
        show_error_dialog(f"Ошибка сохранения в Excel: {str(e)}")
        i, j = user_data
        dpg.configure_item(item=f'load_{i}_{j}', show=False)

def del_lin_graph(sender, app_data, user_data):
    global lin_graphs, lin_series

    i, j = user_data

    dpg.delete_item(item=f'lin_graph_{user_data[0]}_{user_data[1]}')
    lin_graphs.remove(user_data)

    print(dpg.does_item_exist(f'series_{user_data[0]}_{user_data[1]}'))
    if j in lin_series[i] and lin_series[i][j] is not None:
        if dpg.does_item_exist(lin_series[i][j]):
            dpg.delete_item(lin_series[i][j])

def del_field(sender, app_data, user_data):
    global lin_field

    i = user_data

    dpg.delete_item(item=f'lin_field_{i}')
    lin_fields.remove(i)

def show_grex(sender, app_data, user_data):
    if app_data:
        dpg.configure_item(item='grex_group', show=True)
    else:
        dpg.configure_item(item='grex_group', show=False)

def show_about_callback():
    # Размеры viewport (клиентская область)
    vp_width = dpg.get_viewport_client_width()
    vp_height = dpg.get_viewport_client_height()

    # Размеры окна "about_window" (можно взять из задания или получить динамически)
    win_width = 600  # как задано при создании
    win_height = 600  # как задано при создании

    # Координаты для центрирования
    pos_x = (vp_width - win_width) // 2
    pos_y = (vp_height - win_height) // 2

    # Устанавливаем позицию и показываем окно
    dpg.set_item_pos("about_window", [pos_x, pos_y])
    dpg.show_item("about_window")

dpg.create_context()



# загружаем шрифты
if True:
    big_let_start = 0x00C0  # Capital "A" in cyrillic alphabet
    big_let_end = 0x00DF  # Capital "Я" in cyrillic alphabet
    small_let_end = 0x00FF  # small "я" in cyrillic alphabet
    remap_big_let = 0x0410  # Starting number for remapped cyrillic alphabet
    alph_len = big_let_end - big_let_start + 1  # adds the shift from big letters to small
    alph_shift = remap_big_let - big_let_start  # adds the shift from remapped to non-remapped
    with dpg.font_registry():
        with dpg.font("fonts/Roboto_Condensed-Regular.ttf", 20) as default_font:
            dpg.add_font_range_hint(dpg.mvFontRangeHint_Default)
            dpg.add_font_range_hint(dpg.mvFontRangeHint_Cyrillic)
            biglet = remap_big_let  # Starting number for remapped cyrillic alphabet
            for i1 in range(big_let_start, big_let_end + 1):  # Cycle through big letters in cyrillic alphabet
                dpg.add_char_remap(i1, biglet)  # Remap the big cyrillic letter
                dpg.add_char_remap(i1 + alph_len, biglet + alph_len)  # Remap the small cyrillic letter
                biglet += 1  # choose next letter
            dpg.bind_font(default_font)

dpg.create_viewport(title='Transformer', width=1400, height=850)

with dpg.window(tag='win1', width=550, height=350):
    with dpg.table(header_row=False, width=-1):
        dpg.add_table_column(width_stretch=True)  # левая колонка занимает всё свободное место
        dpg.add_table_column(width_fixed=True, init_width_or_weight=120)  # правая под кнопку
        with dpg.table_row():
            dpg.add_text("")  # пустота слева
            dpg.add_button(label="О программе", width=120, callback=show_about_callback)

    with dpg.tab_bar(tag='tab_bar1'):
        with dpg.tab(label='Загрузка данных', tag='tab_1'):
            with dpg.group(horizontal=True):
                dpg.add_button(label='Загрузить таблицы', tag='btn_table', callback=load_to_db)
                dpg.add_button(label='Загрузить файл базы данных', tag='btn_db', callback=load_to_db)
            dpg.add_input_text(default_value='', tag='input_dir', callback=comma_input_callback, width=-1)
            dpg.add_progress_bar(default_value=0, tag='progress_bar', overlay='', show=False)
            dpg.add_text(tag='text_1', default_value='')
        with dpg.tab(label='Расчет', tag='tab_2'):
                    dpg.add_text(default_value='Каталожные данные трансформатора')
                    with dpg.group(horizontal=True):
                        dpg.add_text(default_value='S, МВА =')
                        dpg.add_input_text(tag='S', width=100, default_value='2,5', callback=comma_input_callback)
                    with dpg.group(horizontal=True):
                        dpg.add_text(default_value='Uном вн, кВ =')
                        dpg.add_input_text(tag='U vn', width=100, default_value='6', callback=comma_input_callback)
                    with dpg.group(horizontal=True):
                        dpg.add_text(default_value='Uном нн, кВ =')
                        dpg.add_input_text(tag='U nn', width=100, default_value='0,4', callback=comma_input_callback)
                    with dpg.group(horizontal=True):
                        dpg.add_text(default_value='Uк, % =')
                        dpg.add_input_text(tag='U k', width=100, default_value='7', callback=comma_input_callback)
                    with dpg.group(horizontal=True):
                        dpg.add_text(default_value='dPк, кВт =')
                        dpg.add_input_text(tag='dPk', width=100, default_value='20,7', callback=comma_input_callback)
                    with dpg.group(horizontal=True):
                        dpg.add_text(default_value='dPx, кВт =')
                        dpg.add_input_text(tag='Phh', width=100, default_value='5,04', callback=comma_input_callback)
                    with dpg.group(horizontal=True):
                        dpg.add_text(default_value='Ix, % =')
                        dpg.add_input_text(tag='Ihh', width=100, default_value='0,8', callback=comma_input_callback)
                    dpg.add_separator()
                    with dpg.group(horizontal=True):
                        dpg.add_text(default_value='Расчетные данные трансформатора')
                        dpg.add_button(label='Рассчитать', tag='calc_trans_btn', callback=calc_trans)

                    with dpg.group(horizontal=True):
                        dpg.add_text(default_value='dQх, квар =')
                        dpg.add_input_text(tag='dQ', width=100, callback=comma_input_callback)
                    with dpg.group(horizontal=True):
                        dpg.add_text(default_value='Rт, Ом =')
                        dpg.add_input_text(tag='Rt', width=100, callback=comma_input_callback)
                    with dpg.group(horizontal=True):
                        dpg.add_text(default_value='Xт, Ом =')
                        dpg.add_input_text(tag='Xt', width=100, callback=comma_input_callback)
                    with dpg.group(horizontal=True):
                        dpg.add_text(default_value='Gт, мкСм =')
                        dpg.add_input_text(tag='Gt', width=100, callback=comma_input_callback)
                    with dpg.group(horizontal=True):
                        dpg.add_text(default_value='Bт, мкСм =')
                        dpg.add_input_text(tag='Bt', width=100, callback=comma_input_callback)
                    with dpg.group(horizontal=True):
                        dpg.add_text(default_value='Kт =')
                        dpg.add_input_text(tag='kt', width=100, callback=comma_input_callback)
                    dpg.add_separator()
                    dpg.add_text(default_value='Коэффициенты трансформации ИТТ, ИТН')

                    with dpg.group(horizontal=True):
                        dpg.add_text(default_value='Кт вн =')
                        dpg.add_input_text(tag='kt vn', width=100, default_value='60', callback=comma_input_callback)

                    with dpg.group(horizontal=True):
                        dpg.add_text(default_value='Кн вн =')
                        dpg.add_input_text(tag='kn vn', width=100, default_value='60', callback=comma_input_callback)

                    with dpg.group(horizontal=True):
                        dpg.add_text(default_value='Кт нн =')
                        dpg.add_input_text(tag='kt nn', width=100, default_value='800', callback=comma_input_callback)

                    with dpg.group(horizontal=True):
                        dpg.add_text(default_value='Кн нн =')
                        dpg.add_input_text(tag='kn nn', width=100, default_value='1', callback=comma_input_callback)

                    dpg.add_separator()
                    with dpg.group(horizontal=True):
                        dpg.add_text(default_value='Учитывать коэффициенты систематической погрешности ТН')
                        dpg.add_checkbox(tag='koff', default_value=True, callback=show_grex, user_data='koff')
                    with dpg.group(horizontal=True, tag='grex_group'):
                        dpg.add_text(default_value='a =')
                        dpg.add_input_text(tag='kn vn a', width=100, default_value='0,0044', callback=comma_input_callback)

                        dpg.add_text(default_value='w =')
                        dpg.add_input_text(tag='kn vn w', width=100, default_value='0,01467', callback=comma_input_callback)
                    dpg.add_separator()
                    with dpg.group(horizontal=True):
                        dpg.add_button(label='Рассчитать и загрузить в базу данных', tag='calculate_btn', callback=calculate)
                        dpg.add_button(label='Загрузить файл базы данных вычислений', tag='load_calc_btn', callback=load_calc_file)
                    dpg.add_progress_bar(default_value=0, tag='progress_bar_2', show=False)
                    dpg.add_text(tag='text_2', default_value='')

                    with dpg.group(horizontal=True):
                        dpg.add_button(label='Добавить поле точечных графиков', tag='add_lin_graph_btn', show=False, callback=add_lin_plot)

with dpg.window(label="О программе", tag="about_window", width=600, height=600, show=False):
    dpg.add_text("Transformer 3.0")
    dpg.add_text("17.03.2026")
    dpg.add_separator()
    dpg.add_text("Программа разработана в рамках выпускной квалификационной работы на тему 'Оценивание параметров трехфазных двухобмоточных трансформаторов по данным синхронизированных измерений'", wrap=600)
    dpg.add_text(
        "ПО позволяет осуществлять агрегацию данных с устройств синхронизированных векторных измерений, расположенных на трансформаторе, производить расчет параметров данного трансформатора и производить построение графиков на основе полученных значений", wrap=600)
    dpg.add_separator()
    dpg.add_text("Правообладатели программы: ")
    dpg.add_text("Максименко Яна Александровна, Северо-Кавказский федеральный университет", wrap=600)
    dpg.add_separator()
    dpg.add_text("Соавторы программы:")
    dpg.add_text("Максименко Яна Александровна - MaximenkoYAA@yandex.ru")
    dpg.add_text("Кононов Юрий Григорьевич - iukononov@ncfu.ru")
    dpg.add_text("Звада Павел Александрович")
    dpg.add_text("Овчаренко Александр Витальевич")
    dpg.add_text("Арчебасов Владислав Юрьевич")
    dpg.add_text("Мартусенко Виталий Евгеньевич")
    dpg.add_text("Мхце Ренат Казбекович")





dpg.set_primary_window('win1', True)
dpg.setup_dearpygui()
dpg.show_viewport()
dpg.start_dearpygui()
dpg.destroy_context()